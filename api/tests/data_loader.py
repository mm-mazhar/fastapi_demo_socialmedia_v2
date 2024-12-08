from typing import Any, List, Optional

import openpyxl


def get_data_from_excel(
    filename: str, sheet_name: str, column_names: List[str]
) -> Optional[List[List[str]]]:
    print("*" * 120)
    print(f"Loading test data from Excel File......")
    print(f"Test Data File: {filename}, Sheet: {sheet_name}, Columns: {column_names}")

    workbook: openpyxl = openpyxl.load_workbook(filename, data_only=True)
    sheet: Any = (
        workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
    )

    # Initialize an empty list to hold data for each specified column
    data:list[list] = [[] for _ in column_names]

    # Read the header row to get column indices
    headers: Any = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    column_indices: list[Any | None] = [
        headers.index(column_name) if column_name in headers else None
        for column_name in column_names
    ]

    # Iterate over the rows, starting from row 2 to skip the header
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for i, column_index in enumerate(column_indices):
            # Check if the column index is valid and the row has enough columns
            if column_index is not None and len(row) > column_index:
                data[i].append(row[column_index])
            else:
                # Append None for rows that are shorter than the column_index or if column_name not found
                data[i].append(None)

    if all(not column_data for column_data in data):
        print(
            f"No data found in the specified columns '{column_names}' of the sheet '{sheet_name}'."
        )
        print("*" * 120)
        return None

    print("*" * 120)
    return data
